import os
import xlrd
import os.path
import sys
import datetime
import time
import logging
from email import emailtask, emailerror


class pathnames(object):
    # sharedfile = os.path.dirname(__file__)  # This is for testing only. Afterwards, I will change this to a shared file on a server.
    sharedfile = 'G:\\Quality\\Experience_and_Performance_Improvement\\Programs & Projects\\Lean Portfolio\\Network\\Projects\\CEE 5S Break Room'
    excelfile = '5S_Volunteers.xlsx'
    logfile = 'Logfile.txt'
    EXCELFILEPATH = os.path.join(sharedfile, excelfile)
    LOGPATH = os.path.join(sharedfile, logfile)


def BreakRoomEmailString(ls_br_names):
    '''
    Takes a list of employee names and inserts them into a text script which will become the body of the break room email
    @ ls_br_names (list) - list of employee names in order of primary [0] then secondary [1]
    '''
    primary = ls_br_names[0]
    backup = ls_br_names[1]
    msg = 'Hello ' + primary + ',' + '\n' + 'You have been tasked with the break room 5S checklist this week and ' + backup +' will be your backup.' + '\n' + 'Thank you,' + '\n' + 'Your automated task reminder :)'
    return msg

def CopyRoomEmailString(ls_cr_names):
    '''
    Takes a list of employee names and inserts them into a text script which will become the body of the copy room email
    @ ls_cr_names (list) - list of employee names in order of primary [0] then secondary [1]
    '''
    primary = ls_cr_names[0]
    backup = ls_cr_names[1]
    msg = 'Hello ' + primary + ',' + '\n' + 'You have been tasked with the copy room 5S checklist this week and ' + backup +' will be your backup.' + '\n' + 'Thank you,' + '\n' + 'Your automated task reminder :)'
    return msg

class c_dataobj(object):

    def __init__(self):

        logging.basicConfig(filemode = 'a', format='%(levelname)s:%(asctime)s:%(message)s', datefmt='%m/%d/%Y %I:%M:%S %p', filename= pathnames.LOGPATH, level=logging.INFO)
        logging.info('\n')
        logging.info('*****************Program Start**********************')

    def excel_data(self):
        try:
            book = xlrd.open_workbook(pathnames.EXCELFILEPATH)
            datesheet = book.sheet_by_index(0)
            idsheet = book.sheet_by_index(1)
        except:
            logging.error('Error locating one of the input workbooks.')
            title = 'ACTION REQUIRED: Error in the Automated email Reminder System!'
            msg = 'There was an error loading the 5S volunteer excel workbook'
            email.erroremail(title, msg)
            sys.exit(0)

        # today = datetime.datetime.now() - datetime.timedelta(1)
        today = datetime.datetime.now()
        today = today.replace(hour=0, minute=0, second=0, microsecond=0)
        # print(today)

        for row in range(datesheet.nrows):
            date = datesheet.cell(row, 0).value
            if (isinstance(date, float)):
                reminder_date = datetime.datetime(*xlrd.xldate_as_tuple(date, book.datemode))
                if reminder_date == today:
                    br_id_number = int(datesheet.cell(row, 1).value)
                    cr_id_number = int(datesheet.cell(row, 2).value)
                    ids_left = datesheet.nrows - row
                    c_dataobj.write_excel_date(row)
                    if ids_left <= 2:
                        title = 'ACTION REQUIRED: Error in the Automated email Reminder System!!'
                        msg = 'The breakroom / copyroom task reminder workbook needs to be updated. There are only a few dates left.' + '\n' + str(pathnames.EXCELFILEPATH)
                        email.erroremail(title, msg)
                        break
                # elif (datesheet.cell(row, 3).value == ""):
                #     last_sent = datetime.datetime(*xlrd.xldate_as_tuple(datesheet.cell(row - 1, 3).value, book.datemode))
                #     should_last_sent = today - datetime.timedelta(7)
                #     if should_last_sent > last_sent:
                #         print('Whoa')
                #     else:
                #         print('OK')

                    # print(datesheet.cell(row-1,3).value)
                    # print(reminder_date)
                    # print(today)
                    # print(today - reminder_date)
                    # sys.exit(0)
                    # do something when today - reminder_date is greater than 7
        # time.sleep(5)
        # sys.exit(0)

        ls_br_names = []
        ls_br_emails = []
        ls_cr_names = []
        ls_cr_emails = []

        try:
            if br_id_number:
                for row in range(idsheet.nrows):
                    matching_id_number = idsheet.cell(row, 0).value
                    if br_id_number == matching_id_number:
                        ls_br_names.append(idsheet.cell(row, 1).value)
                        ls_br_emails.append(idsheet.cell(row, 2).value)
                        try:
                            backup_name = idsheet.cell(row+1, 1).value
                            backup_email = idsheet.cell(row+1, 2). value
                        except:
                            backup_name = idsheet.cell(0,1).value
                            backup_email = idsheet.cell(0,2).value
                        ls_br_names.append(backup_name)
                        ls_br_emails.append(backup_email)

                    elif cr_id_number == matching_id_number:
                        ls_cr_names.append(idsheet.cell(row, 1).value)
                        ls_cr_emails.append(idsheet.cell(row, 2).value)
                        try:
                            backup_name = idsheet.cell(row+1, 1).value
                            backup_email = idsheet.cell(row+1, 2). value
                        except:
                            backup_name = idsheet.cell(0,1).value
                            backup_email = idsheet.cell(0,2).value
                        ls_cr_names.append(backup_name)
                        ls_cr_emails.append(backup_email)

        except:
            print('did not match the date. Terminating program')
            logging.info('Today is not Monday')
            sys.exit(0)

        logging.info('\n')
        logging.info('*****************Program Start**********************')

        c_dataobj.notifyemail(ls_br_names, ls_br_emails, ls_cr_names, ls_cr_emails)

    def write_excel_date(n_row):
        import openpyxl
        wb = openpyxl.load_workbook(filename= pathnames.EXCELFILEPATH)
        ws = wb['Schedule']

        o_today = datetime.datetime.now()
        # today = today.replace(hour=0, minute=0, second=0, microsecond=0)
        s_today = o_today.strftime('%m/%d/%Y')
        ws.cell(row= n_row+1, column=4).value = s_today
        wb.save(pathnames.EXCELFILEPATH)


    def notifyemail(ls_br_names, ls_br_emails, ls_cr_names, ls_cr_emails):

        br_email_msg = BreakRoomEmailString(ls_br_names)
        cr_email_msg = CopyRoomEmailString(ls_cr_names)

        emailtask(br_email_msg, ls_br_emails)
        logging.info(str(ls_br_names[0]) + ' and ' + str(ls_br_names[1]) + ' have been notified about the break room 5S checklist')
        emailtask(cr_email_msg, ls_cr_emails)
        logging.info(str(ls_cr_names[0]) + ' and ' + str(ls_cr_names[1]) + ' have been notified about the copy room 5S checklist')

if __name__ == '__main__':

    my_obj = c_dataobj()
    print('running...')
    # time.sleep(5)

    my_obj.excel_data()


